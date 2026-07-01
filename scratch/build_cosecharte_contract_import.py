import json
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

contracts_path = Path(r'C:\Users\ASUS\OneDrive\Escritorio\1. MATRIZ VENCIMIENTO DE CONTRATOS (1).xlsx')
employees_path = Path(r'C:\Users\ASUS\Programacion IA\krh-recursoshumanos\EmpleadosCosecharte.xlsx')
out_path = Path(r'C:\Users\ASUS\Programacion IA\krh-recursoshumanos\scratch\cosecharte_contract_import.json')

def norm_doc(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return ''.join(ch for ch in str(v).strip() if ch.isdigit())

def iso(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()

def clean_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None

# Salary map from employee source workbook.
ewb = load_workbook(employees_path, data_only=True, read_only=True)
ews = ewb['Hoja2']
headers = [str(c).strip() if c is not None else '' for c in next(ews.iter_rows(min_row=1, max_row=1, values_only=True))]
idx_doc = headers.index('Identidificación') if 'Identidificación' in headers else headers.index('Identidificaci�n') if 'Identidificaci�n' in headers else 10
idx_salary = headers.index('salario')
salaries = {}
for row in ews.iter_rows(min_row=2, values_only=True):
    doc = norm_doc(row[idx_doc])
    if not doc:
        continue
    salary = row[idx_salary]
    try:
        salary = float(salary) if salary is not None else None
    except Exception:
        salary = None
    if salary and salary > 0:
        salaries[doc] = salary

cwb = load_workbook(contracts_path, data_only=True, read_only=True)
ws = cwb['A TERMINO FIJO']
rows = []
missing_salary = []
for rn, row in enumerate(ws.iter_rows(min_row=3, max_row=167, min_col=1, max_col=11, values_only=True), start=3):
    doc = norm_doc(row[0])
    if not doc:
        continue
    salary = salaries.get(doc)
    if not salary:
        missing_salary.append({'row': rn, 'document_number': doc})
    rows.append({
        'excel_row': rn,
        'document_number': doc,
        'contract_type': 'fijo',
        'source_contract_type': clean_text(row[1]),
        'operation_center_name': clean_text(row[2]),
        'salary_type': 'mensual',
        'source_payment_mode': clean_text(row[3]),
        'start_date': iso(row[4]),
        'duration_months': int(row[5]) if isinstance(row[5], (int, float)) and row[5] else None,
        'end_date': iso(row[6]),
        'extension_end_dates': [iso(row[i]) for i in (7,8,9) if iso(row[i])],
        'salary': salary,
        'special_clauses': clean_text(row[10]),
    })

out_path.parent.mkdir(exist_ok=True)
out_path.write_text(json.dumps(rows, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
print(json.dumps({
    'contracts_rows': len(rows),
    'salary_rows': len(salaries),
    'missing_salary_count': len(missing_salary),
    'missing_salary_sample': missing_salary[:10],
    'output': str(out_path),
}, ensure_ascii=False, indent=2))
